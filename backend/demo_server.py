"""
Simplified demo server - Ryan Serhant Response Tool
Shows the actual API endpoints and response format without external dependencies
"""

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import JSONResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import json
from datetime import datetime
from typing import Optional
import os
import tempfile

# Using Groq (FREE) for response generation
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

app = FastAPI(title="Ryan Serhant Response Tool", version="1.0.0")

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Sample responses database
SAMPLE_RESPONSES = {
    "price": {
        "response_text": """Listen, price objections are the easiest ones to overcome because they're not really about price—they're about VALUE. Here's what you do:

First, never defend the price. That's weakness. Instead, ask: "What specifically concerns you about the investment?" Nine times out of ten, they don't actually understand what they're buying or what you're bringing to the table.

Then—and this is critical—reframe the conversation. Don't talk about what the property COSTS. Talk about what it MAKES or SAVES or PROVIDES. You're not selling a house, you're selling their future. You're selling peace of mind. You're selling the lifestyle they want.

Here's the script I use: "I hear you on price. What I'm really asking is—are we talking about the same property? Because if you saw someone else sell something similar for less, that's actually concerning to me. Either they're missing something you need, or we need to talk about what makes THIS property worth the premium."

Then be quiet. Let them talk. Most agents jump in and start justifying. Don't do that. The person who talks first loses.

One more thing: sometimes price IS the real issue—they legitimately can't afford it or the market won't bear it. That's okay. You pivot to creative solutions: seller financing, lease options, staging to increase perceived value. But you never, ever apologize for quality.

The bottom line: price objections mean they're interested. No interest = no objections. So treat it like the opportunity it is.""",
        "citations": [
            {
                "rank": 1,
                "source": "book",
                "title": "Sell It Like Serhant - Negotiation Chapter",
                "quality_score": 94,
                "excerpt": "Price objections are not about price—they're about value perception. The key is to reframe the conversation from cost to benefit.",
                "url": "https://example.com/book/negotiation",
                "topics": ["objection_handling", "negotiation"]
            },
            {
                "rank": 2,
                "source": "course",
                "title": "Advanced Selling - Lesson 4",
                "quality_score": 90,
                "excerpt": "Use the silence strategy after presenting your value. The person who speaks first loses the negotiation.",
                "url": "https://example.com/course/lesson-4",
                "topics": ["scripts", "objection_handling"]
            }
        ]
    },
    "closing": {
        "response_text": """Closing is about confidence and clarity. Here's what most agents miss: they ask for the deal but leave room for the buyer to say no. That's a mistake.

When you're ready to close, you need to be direct. Not aggressive—direct. Say: "Let's move forward. Here's what happens next..." Then tell them exactly what the next steps are.

The key is making it easy for them to say yes. Remove friction. Remove options. Give them ONE path forward.

And here's the psychological play: when you present the deal with confidence and clarity, the buyer feels that confidence. They trust you more. They're more likely to commit.

But the moment you hedge or ask permission, they sense it. They know they can push back. So don't give them that opening.

Own the close. Own the conversation. Own the outcome.""",
        "citations": [
            {
                "rank": 1,
                "source": "course",
                "title": "Advanced Selling - Lesson 7: Closing Techniques",
                "quality_score": 92,
                "excerpt": "Closing is about confidence. Present the deal with certainty and the buyer will match your energy.",
                "url": "https://example.com/course/lesson-7",
                "topics": ["closing", "personal_brand"]
            }
        ]
    }
}


@app.get("/api/health")
async def health():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "Ryan Serhant Response Tool"
    }


# ============================================================
# THE ROUNDTABLE - selectable expert influences
# ============================================================
try:
    from backend.experts_kb import EXPERTS, DEFAULT_EXPERT_IDS, DEEP_PANEL_MAX
    from backend.kb_retrieval import retrieve_kb_sections
except ImportError:
    from experts_kb import EXPERTS, DEFAULT_EXPERT_IDS, DEEP_PANEL_MAX
    from kb_retrieval import retrieve_kb_sections


def build_system_prompt(expert_ids: Optional[list] = None) -> str:
    """Focused panels (<= DEEP_PANEL_MAX advisors) get FULL knowledge dossiers
    (frameworks, named systems, methodology). Broad panels get one-line style
    fingerprints to stay within free-tier token limits."""
    ids = [e for e in (expert_ids or []) if e in EXPERTS] or DEFAULT_EXPERT_IDS
    if len(ids) <= DEEP_PANEL_MAX:
        panel = "\n\n".join(
            f"### {EXPERTS[e]['name']} - {EXPERTS[e]['focus']}\n"
            f"Voice: {EXPERTS[e]['style']}\n"
            f"Knowledge & frameworks: {EXPERTS[e]['deep']}"
            for e in ids
        )
    else:
        panel = "\n".join(
            f"- {EXPERTS[e]['name']} ({EXPERTS[e]['focus']}): {EXPERTS[e]['style']}"
            for e in ids
        )
    return BASE_SYSTEM_PROMPT.replace("{PANEL}", panel)


BASE_SYSTEM_PROMPT = """You are The Roundtable - an elite AI real estate coach that synthesizes the collective wisdom of the industry's best minds into one clear, direct advisor. The agent talking to you chose which expert perspectives sit at their table today.

ACTIVE ADVISOR PERSPECTIVES AT THE TABLE:
{PANEL}

YOUR VOICE:
- Direct, confident, warm. You talk like a trusted mentor who happens to be a killer closer.
- Short punchy sentences. You ask questions back. You reference specifics.
- You NEVER give generic advice. Every response is tailored to exactly what they showed you or asked you.
- Blend the active perspectives into ONE coherent answer. When the perspectives genuinely differ, briefly show the contrast (e.g., "The prospecting play here is X. The relationship play is Y - pick based on your timeline.").
- Draw on the frameworks of the active advisors when they genuinely fit; never name-drop for its own sake.
- ESPANOL: If the agent writes in Spanish or asks for Spanish output, respond in fully natural, professional Spanish (US Latino register - Southern California real estate Spanish). When they mention Spanish-speaking clients or bilingual farms, proactively offer the Spanish version of any script, text, or letter alongside the English.

WHAT YOU CAN REVIEW — and how:
- PRELIMINARY TITLE REPORTS / PROPERTY PROFILES: Walk through what you see. Flag liens, deeds of trust, easements, tax status, vesting issues, ownership history. Tell them what to verify before listing or writing an offer, and what talking points this gives them with the seller.
- LISTING PRESENTATIONS & FLYERS: Critique like a marketer. Headline, hero image, story, call to action, what's missing, what to cut. Be specific about THEIR content.
- HEADSHOTS & PHOTOS: Give honest branding feedback — lighting, wardrobe, expression, background, what it signals to a luxury vs. first-time buyer.
- CALL SCRIPTS: Rewrite lines. Show them the exact words. Point out where they'd lose the person on the phone.
- EMAILS & TEXTS: Rewrite for punch. Subject lines, first sentence hooks, clear ask.
- LEAD LISTS / SPREADSHEETS: Look at the actual segments and data. Suggest an attack plan — who to call first, what angle per segment.
- CONTRACTS/OFFERS: Flag the business terms that matter (you're not their lawyer — say so once, briefly).

RULES:
1. ALWAYS anchor to the actual content given to you. Quote it, name names, cite the address, reference the specific lien or the specific headline. Prove you read it.
2. If they gave you a document AND a question, answer THEIR question first, using the document.
3. If the document is partial or garbled, work with what IS there — never complain about formatting.
4. End with either one sharp question back to them or one concrete next action. Not both, not a list of five.
5. Never say "As an AI." You are their coach - The Roundtable - in this conversation.
6. Length: match the task. A headshot critique might be 150 words. A prelim review might be 500. Don't pad."""


async def generate_response_with_groq(query: str, style: str, context: Optional[str] = None,
                                      history: Optional[list] = None,
                                      expert_ids: Optional[list] = None) -> Optional[str]:
    """Generate response using Groq API (FREE - 10x faster than Claude!)

    history: prior conversation turns [{role: 'user'|'assistant', content: str}, ...]
    expert_ids: which advisor influences the client selected for their panel."""
    try:
        groq_api_key = os.getenv("GROQ_API_KEY")
        if not groq_api_key:
            return None

        client = Groq(api_key=groq_api_key)

        if context:
            user_content = f"""Here is the document/content I want you to review:

---BEGIN DOCUMENT---
{context}
---END DOCUMENT---

My question/request: {query if query else 'Review this and give me your take — what matters, what to watch out for, and what I should do with it.'}"""
        else:
            user_content = query

        # Topic retrieval: inject the most relevant research-library sections
        # (works even when the panel is too large for full dossiers)
        retrieval_probe = query + " " + (context[:1500] if context else "")
        kb_extra = retrieve_kb_sections(retrieval_probe)
        system_content = build_system_prompt(expert_ids)
        if kb_extra:
            system_content += ("\n\nRESEARCH LIBRARY EXCERPTS (deep sourced material relevant to this "
                               "question - draw on it where it fits):\n" + kb_extra)

        messages = [{"role": "system", "content": system_content}]
        # Replay prior turns (validated + capped) so follow-up replies keep full context
        if history:
            for turn in history[-12:]:
                role = turn.get("role")
                content = str(turn.get("content", ""))[:8000]
                if role in ("user", "assistant") and content:
                    messages.append({"role": role, "content": content})
        messages.append({"role": "user", "content": user_content})

        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",  # Free, fast, and powerful
            messages=messages,
            max_tokens=1200,
            temperature=0.7
        )

        return response.choices[0].message.content

    except Exception as e:
        print(f"Groq error: {e}")
        return None


# Groq's FREE vision model - used for OCR on scanned PDFs and real photo reviews
VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"


def _groq_client() -> Optional["Groq"]:
    key = os.getenv("GROQ_API_KEY")
    if not key or not GROQ_AVAILABLE:
        return None
    return Groq(api_key=key)


def ocr_pdf_with_groq_vision(file_path: str, max_pages: int = 8) -> str:
    """OCR a scanned (image-only) PDF: render pages to images with PyMuPDF,
    transcribe each with Groq's free vision model. $0 cost."""
    import base64
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return ""
    client = _groq_client()
    if client is None:
        return ""
    try:
        doc = fitz.open(file_path)
    except Exception:
        return ""
    import time
    total = doc.page_count
    pages = min(total, max_pages)
    out = []
    for i in range(pages):
        pix = doc[i].get_pixmap(dpi=100)
        b64 = base64.b64encode(pix.tobytes("png")).decode()
        # Free tier = 30k tokens/min on the vision model; retry with a pause on 429s
        for attempt in range(3):
            try:
                r = client.chat.completions.create(
                    model=VISION_MODEL,
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Transcribe ALL text visible in this document page, exactly as written. Render tables as plain text rows. Output ONLY the transcription, no commentary."},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                        ],
                    }],
                    max_tokens=2000,
                    temperature=0,
                )
                out.append(f"--- Page {i + 1} ---\n{r.choices[0].message.content}")
                break
            except Exception as e:
                if "rate_limit" in str(e) and attempt < 2:
                    time.sleep(8)
                    continue
                print(f"Vision OCR error on page {i + 1}: {e}")
                break
    doc.close()
    text = "\n\n".join(out)
    if total > pages and text:
        text += f"\n\n[Note: this document has {total} pages; the first {pages} were read.]"
    return text


def analyze_image_with_groq(file_path: str, query: Optional[str], file_name: str,
                            expert_ids: Optional[list] = None) -> Optional[str]:
    """The panel actually LOOKS at an uploaded photo (headshot, flyer, listing photo,
    room to stage) and critiques it using the vision model."""
    import base64
    client = _groq_client()
    if client is None:
        return None
    ext = os.path.splitext(file_name)[1].lower().lstrip(".")
    mime = "jpeg" if ext in ("jpg", "jpeg") else "png"
    try:
        with open(file_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        user_text = query or "Review this image and give me your honest take - what works, what doesn't, and what I should change."
        r = client.chat.completions.create(
            model=VISION_MODEL,
            messages=[
                {"role": "system", "content": build_system_prompt(expert_ids)},
                {"role": "user", "content": [
                    {"type": "text", "text": f"(Uploaded image: {file_name})\n\n{user_text}"},
                    {"type": "image_url", "image_url": {"url": f"data:image/{mime};base64,{b64}"}},
                ]},
            ],
            max_tokens=1000,
            temperature=0.7,
        )
        return r.choices[0].message.content
    except Exception as e:
        print(f"Vision analyze error: {e}")
        return None


def _nlm():
    try:
        from backend import nlm_bridge
    except ImportError:
        import nlm_bridge
    return nlm_bridge


@app.get("/api/nlm-ready")
async def nlm_ready_endpoint():
    """Is real NotebookLM audio generation configured on this server?"""
    try:
        return {"ready": _nlm().nlm_ready()}
    except Exception:
        return {"ready": False}


@app.post("/api/nlm-audio")
async def nlm_audio_start(data: dict):
    """Start a REAL NotebookLM Audio Overview job for an episode source doc."""
    bridge = _nlm()
    if not bridge.nlm_ready():
        return JSONResponse(status_code=501, content={
            "status": "error",
            "message": "NotebookLM auth isn't configured on this server yet."})
    title = (data.get("title") or "Roundtable Episode")[:120]
    source_doc = (data.get("source_doc") or "")[:200_000]
    if len(source_doc) < 200:
        return JSONResponse(status_code=400, content={"status": "error", "message": "source_doc required"})
    job_id = bridge.start_audio_job(title, source_doc, data.get("language", "en"),
                                    (data.get("instructions") or "")[:500] or None)
    return {"status": "started", "job_id": job_id}


@app.get("/api/nlm-audio/{job_id}")
async def nlm_audio_status(job_id: str):
    s = _nlm().job_status(job_id)
    if s is None:
        return JSONResponse(status_code=404, content={"status": "error", "message": "unknown job"})
    return s


@app.get("/api/nlm-audio/{job_id}/file")
async def nlm_audio_file(job_id: str):
    from fastapi.responses import FileResponse
    f = _nlm().job_file(job_id)
    if not f:
        return JSONResponse(status_code=404, content={"status": "error", "message": "no file yet"})
    return FileResponse(f, media_type="audio/mpeg", filename="roundtable_episode.mp3")


try:
    from backend.podcast_concepts import PODCAST_ADVISORS
    from backend.kb_retrieval import get_kb
except ImportError:
    from podcast_concepts import PODCAST_ADVISORS
    from kb_retrieval import get_kb


@app.get("/api/podcast-concepts")
async def podcast_concepts():
    """The advisor episode catalog: first names, voice genders, and 3 researched
    episode concepts per advisor."""
    out = []
    for aid, meta in PODCAST_ADVISORS.items():
        e = EXPERTS.get(aid, {})
        out.append({"id": aid, "name": e.get("name", meta["first"]),
                    "first": meta["first"], "gender": meta["gender"],
                    "focus": e.get("focus", ""), "concepts": meta["concepts"]})
    return {"advisors": out}


@app.post("/api/podcast-script")
async def podcast_script(data: dict):
    """Podcast Studio: topic or advisor-persona episodes.

    advisors: [] -> default ALEX/MORGAN show grounded via retrieval.
    advisors: [id] -> interview: ALEX interviews that advisor (their first name,
              their researched ideology).
    advisors: [id1, id2] -> duo show: both hosts ARE the advisors, speaking their
              own researched philosophies, agreeing and clashing where they truly differ.
    """
    topic = (data.get("topic") or "real estate success habits").strip()[:200]
    minutes = int(data.get("minutes") or 6)
    language = data.get("language", "en")
    advisor_ids = [a for a in (data.get("advisors") or []) if a in PODCAST_ADVISORS][:2]

    lang_rule = ("Write the dialogue in natural US-Latino professional Spanish."
                 if language == "es" else "Write the dialogue in English.")

    if advisor_ids:
        metas = [PODCAST_ADVISORS[a] for a in advisor_ids]
        if len(advisor_ids) == 2:
            hosts = [{"name": metas[0]["first"].upper(), "gender": metas[0]["gender"]},
                     {"name": metas[1]["first"].upper(), "gender": metas[1]["gender"]}]
            persona_docs = "\n\n".join(
                f"=== {EXPERTS[a]['name']} (host '{PODCAST_ADVISORS[a]['first'].upper()}') - researched ideology ===\n"
                f"{get_kb(a, 6500)}" for a in advisor_ids)
            cast_rules = (
                f"The two hosts ARE these real advisors, first names only: "
                f"{hosts[0]['name']} and {hosts[1]['name']}. Each host speaks ONLY from their own "
                f"researched ideology below - their frameworks, their numbers, their signature phrases. "
                f"Where their philosophies genuinely differ, let them respectfully CLASH and push back on "
                f"each other before finding the practical middle. No generic co-host filler.")
        else:
            a = advisor_ids[0]
            hosts = [{"name": "ALEX", "gender": "m"},
                     {"name": metas[0]["first"].upper(), "gender": metas[0]["gender"]}]
            persona_docs = (f"=== {EXPERTS[a]['name']} (host '{hosts[1]['name']}') - researched ideology ===\n"
                            f"{get_kb(a, 8000)}")
            cast_rules = (
                f"Host ALEX is the curious interviewer asking what agents are thinking. "
                f"Host {hosts[1]['name']} IS the real advisor - speaking only from the researched ideology "
                f"below: their frameworks, real numbers, and signature phrases, in their voice.")
        research_block = persona_docs
    else:
        hosts = [{"name": "ALEX", "gender": "m"}, {"name": "MORGAN", "gender": "f"}]
        cast_rules = ("Host ALEX (curious, asks the questions agents are thinking, summarizes takeaways). "
                      "Host MORGAN (the expert voice - draws on the research, gives frameworks, numbers, scripts).")
        kb_extra = retrieve_kb_sections(topic, max_chars=6000, max_files=4)
        research_block = kb_extra if kb_extra else "(no library match - solid general coaching knowledge, no invented statistics)"

    host_names = [h["name"] for h in hosts]
    prompt = f"""Create a two-host podcast episode about: {topic}

The show is "The Roundtable Sessions" - a sharp, warm real estate coaching podcast for agents.
{cast_rules}

RESEARCH (ground every claim here - use the real numbers, formulas, named frameworks, and signature phrases):
{research_block}

Rules: ~{minutes} minutes of audio (~{minutes * 140} words). Natural banter like NotebookLM's audio
overviews - short reactions, interruptions, "right", "exactly" - but ALWAYS substantive. Open with a
hook, close with 3 actionable takeaways. Host names in turns must be exactly {host_names[0]} and {host_names[1]}. {lang_rule}

Respond ONLY with JSON:
{{"title": "punchy episode title",
"description": "2-sentence episode description",
"turns": [{{"host": "{host_names[0]}", "text": "..."}}, {{"host": "{host_names[1]}", "text": "..."}}]}}"""

    try:
        client = _groq_client()
        if client is None:
            raise RuntimeError("no client")
        r = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=3000,
            temperature=0.85,
            response_format={"type": "json_object"},
        )
        import json as _json
        ep = _json.loads(r.choices[0].message.content)
        ep["hosts"] = hosts

        # NotebookLM source document: the deep-dive doc an agent can upload for an Audio Overview
        doc_lines = [f"# {ep.get('title', topic)}", "",
                     f"Topic: {topic}", "",
                     ep.get("description", ""), "",
                     "## Episode transcript", ""]
        for t in ep.get("turns", []):
            doc_lines.append(f"**{t.get('host', 'HOST')}:** {t.get('text', '')}")
            doc_lines.append("")
        if research_block and "(no library match" not in research_block:
            doc_lines += ["## Source research (from The Roundtable advisor library)", "", research_block]
        ep["source_doc"] = "\n".join(doc_lines)
        return {"status": "success", "episode": ep}
    except Exception as e:
        print(f"podcast-script error: {e}")
        return JSONResponse(status_code=503, content={
            "status": "error",
            "message": "The panel is waking up - try again in about a minute."})


@app.post("/api/parse-listing")
async def parse_listing_endpoint(data: dict):
    """Flyer Studio importer: paste a Zillow / Redfin / Realtor / MLS link ->
    photos + property facts. Ported from the FlipAI multi-source fallback chain
    (free strategies only)."""
    url = (data.get("url") or "").strip()
    if not url.startswith("http"):
        return JSONResponse(status_code=400, content={"status": "error", "message": "Paste a full listing URL."})
    try:
        from backend.listing_parser import parse_listing
    except ImportError:
        from listing_parser import parse_listing
    import asyncio
    result = await asyncio.to_thread(parse_listing, url)
    return {"status": "success", "listing": result}


@app.post("/api/flyer-copy")
async def flyer_copy(data: dict):
    """Flyer Studio copy engine: property facts in -> polished marketing copy out.
    Uses the marketing-weighted panel (Serhant lifestyle-selling, Glennda storytelling,
    Pantana hooks, Eisen aspiration). Returns strict JSON for the studio to fill."""
    prop = data.get("property", {})
    language = data.get("language", "en")  # en | es | both
    tone = data.get("tone", "luxury")

    facts = "\n".join(f"- {k}: {v}" for k, v in prop.items() if v)
    lang_rule = {
        "en": "Write everything in English.",
        "es": "Write everything in natural US-Latino professional Spanish.",
        "both": "Write every field in English, then add the same fields with '_es' suffix in natural US-Latino professional Spanish.",
    }.get(language, "Write everything in English.")

    prompt = f"""You are the marketing panel of The Roundtable (lifestyle-selling, hook-first, story-driven, aspirational staging language). Write flyer copy for this property:

{facts}

Tone: {tone}. {lang_rule}

Respond with ONLY a JSON object (no markdown, no commentary) with exactly these keys:
{{"headline": "5-8 word emotional hook - sell the lifestyle, not square footage",
"subhead": "one elegant supporting line",
"description": "2-3 flyer-length sentences that make a buyer feel the life they'd live here; end with a soft call to action",
"features": ["4-6 short punchy feature bullets"],
"ig_caption": "an Instagram caption with a hook first line, 2-3 short lines, and 5 relevant hashtags",
"email_subject": "an email subject line under 9 words",
"email_body": "a 3-4 sentence email blast to a buyer database about this listing"}}"""

    try:
        client = _groq_client()
        if client is None:
            raise RuntimeError("no client")
        r = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.8,
            response_format={"type": "json_object"},
        )
        import json as _json
        copy = _json.loads(r.choices[0].message.content)
        return {"status": "success", "copy": copy}
    except Exception as e:
        print(f"flyer-copy error: {e}")
        addr = prop.get("address", "This home")
        return {"status": "fallback", "copy": {
            "headline": f"Welcome Home to {addr}",
            "subhead": "A rare opportunity in a coveted neighborhood",
            "description": "Beautifully presented and move-in ready, this home blends comfort with effortless style. Schedule your private showing today.",
            "features": ["Move-in ready", "Coveted location", "Bright open layout", "Private outdoor space"],
            "ig_caption": f"Just listed: {addr}. DM for a private tour. #justlisted #realestate #dreamhome #newlisting #househunting",
            "email_subject": f"New Listing: {addr}",
            "email_body": f"I just listed {addr} and thought of you. Homes like this rarely last. Reply for a private showing before it hits the broader market.",
        }}


@app.post("/api/localize-script")
async def localize_script(data: dict):
    """Scripts Studio localizer: rewrite a script so it speaks to ONE local market.
    Input: {script, channel, advisor, topic, market:{name, zip, p, sf, dom, cut, inv, mai, rent}}
    The market block comes from the Altos Research snapshot (market_local.json)."""
    script = (data.get("script") or "").strip()
    channel = data.get("channel", "call")
    advisor = data.get("advisor", "")
    topic = data.get("topic", "")
    m = data.get("market") or {}
    if not script or not m.get("name"):
        return {"status": "error", "error": "script and market required"}

    mai = m.get("mai")
    if isinstance(mai, (int, float)):
        temp = ("a strong seller's market - buyers are competing" if mai >= 40 else
                "a seller's market - well-priced homes move" if mai >= 30 else
                "a balanced market - pricing and presentation decide" if mai >= 20 else
                "a buyer's market - sellers need an edge to stand out")
    else:
        temp = "a shifting market"

    stats = []
    if m.get("p"):
        stats.append(f"median list price ${int(m['p']):,}")
    if m.get("dom"):
        stats.append(f"median {int(m['dom'])} days on market")
    if m.get("cut") is not None:
        stats.append(f"{m['cut']}% of active listings have cut their price")
    if m.get("inv") is not None:
        stats.append(f"{m['inv']} active listings")
    if m.get("rent"):
        stats.append(f"median rent ${int(m['rent']):,}/mo")

    chan_rule = {
        "call": "Keep it a natural spoken phone script with the same beats and any agent/owner placeholders intact.",
        "voicemail": "Keep it under 40 seconds spoken (about 90 words), one clear reason to call back.",
        "text": "Keep it 1-3 short sentences, casual-professional, no links, ends with an easy question.",
        "email": "Keep the subject line punchy (under 9 words) and the body under 130 words.",
    }.get(channel, "Keep the same format and length.")

    prompt = f"""You are The Roundtable's script coach. Rewrite this {channel} script so it speaks specifically to the {m['name']} market right now. Current Altos Research data for {m['name']}: {'; '.join(stats)}. Market temperature: {temp}.

Rules:
- Keep the original structure, intent, and advisor voice ({advisor or 'the original voice'}). Topic: {topic or 'general'}.
- Weave in AT MOST two of the local numbers, naturally, the way a sharp local agent quotes stats in conversation - not a data dump.
- Match the tone to the market temperature ({temp}).
- Keep every {{{{placeholder}}}} exactly as written.
- {chan_rule}

ORIGINAL SCRIPT:
{script}

Respond with ONLY a JSON object: {{"script": "the localized script", "why": "one sentence on what you changed for this market"}}"""

    try:
        client = _groq_client()
        if client is None:
            raise RuntimeError("no client")
        r = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.7,
            response_format={"type": "json_object"},
        )
        import json as _json
        out = _json.loads(r.choices[0].message.content)
        if not out.get("script"):
            raise RuntimeError("empty rewrite")
        return {"status": "success", "script": out["script"], "why": out.get("why", "")}
    except Exception as e:
        print(f"localize-script error: {e}")
        return {"status": "error", "error": "The localizer is busy - try again in a few seconds."}


@app.get("/api/experts")
async def list_experts():
    """The full advisor panel available for selection"""
    return {"experts": [{"id": k, "name": v["name"], "focus": v["focus"]} for k, v in EXPERTS.items()]}


@app.get("/api/diag")
async def diag():
    """Diagnostic: is the Groq key loaded and working on this server?"""
    key = os.getenv("GROQ_API_KEY", "")
    info = {
        "groq_sdk_installed": GROQ_AVAILABLE,
        "key_present": bool(key),
        "key_fingerprint": (key[:7] + "..." + key[-4:]) if key else None,
        "key_length": len(key),
        "groq_ping": None,
    }
    try:
        client = _groq_client()
        if client:
            r = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": "say ok"}],
                max_tokens=4,
            )
            info["groq_ping"] = "OK: " + (r.choices[0].message.content or "")
        else:
            info["groq_ping"] = "no client (missing key or sdk)"
    except Exception as e:
        info["groq_ping"] = f"ERROR: {str(e)[:250]}"

    # NotebookLM bridge diagnostics
    nlm_b64 = os.getenv("NLM_STORAGE_B64", "")
    info["nlm_env_present"] = bool(nlm_b64)
    info["nlm_env_length"] = len(nlm_b64)
    if nlm_b64:
        try:
            import base64 as _b64
            import json as _json
            decoded = _b64.b64decode(nlm_b64)
            _json.loads(decoded)
            info["nlm_env_decodes"] = True
        except Exception as e:
            info["nlm_env_decodes"] = f"ERROR: {str(e)[:120]}"
    try:
        info["nlm_ready"] = _nlm().nlm_ready()
    except Exception as e:
        info["nlm_ready"] = f"ERROR: {str(e)[:120]}"
    return info


@app.get("/api/status")
async def status():
    """Get system status and metrics"""
    return {
        "api_status": "online",
        "database_status": "connected",
        "content_stats": {
            "total_content": 95,
            "total_chunks": 8425,
            "total_words": 487230,
            "avg_quality": 82
        },
        "query_stats": {
            "total_queries": 142,
            "avg_latency": 245,
            "cache_hit_rate": 34,
            "success_rate": 100
        }
    }


@app.post("/api/query")
async def query(query_data: dict):
    """
    Main query endpoint - returns response in Ryan's voice with citations
    Using Groq API (FREE - 10x faster than Claude!)

    Request:
    {
        "query": "How do I overcome price objections?",
        "top_k": 5,
        "response_style": "practical",
        "strategy": "hybrid"
    }
    """
    query_text = query_data.get("query", "")
    style = query_data.get("response_style", "practical")
    history = query_data.get("history") or []
    experts = query_data.get("experts") or None  # client-selected advisor panel
    context = query_data.get("context")  # document text carried across the conversation
    if context:
        context = str(context)[:16000]

    # Try to generate response using Groq (FREE!)
    response_text = None
    if GROQ_AVAILABLE:
        response_text = await generate_response_with_groq(query_text, style, context=context,
                                                          history=history, expert_ids=experts)

    # Fallback to sample responses if Groq unavailable
    if not response_text:
        query_text_lower = query_text.lower()
        sample_key = "price" if any(word in query_text_lower for word in ["price", "objection", "overcome"]) else "closing"
        sample = SAMPLE_RESPONSES.get(sample_key, SAMPLE_RESPONSES["price"])
        response_text = sample["response_text"]
        citations = sample["citations"]

        return {
            "status": "success",
            "query": query_data.get("query"),
            "response": {
                "text": response_text,
                "style": style,
                "word_count": len(response_text.split()),
                "reading_time_seconds": len(response_text.split()) // 3,
                "ryan_voice": {
                    "energy": 0.85,
                    "pace_wpm": 145,
                    "confidence": 1.0
                }
            },
            "citations": citations,
            "metadata": {
                "chunks_retrieved": len(citations),
                "confidence_score": 0.89,
                "generation_latency_ms": 245,
                "sources": [c["source"] for c in citations],
                "powered_by": "Groq (FREE!)"
            }
        }

    # Return the REAL Groq-generated response
    return {
        "status": "success",
        "query": query_data.get("query"),
        "response": {
            "text": response_text,
            "style": style,
            "word_count": len(response_text.split()),
            "reading_time_seconds": len(response_text.split()) // 3,
            "ryan_voice": {
                "energy": 0.85,
                "pace_wpm": 145,
                "confidence": 1.0
            }
        },
        "citations": [],
        "metadata": {
            "chunks_retrieved": 0,
            "confidence_score": 0.92,
            "generation_latency_ms": 245,
            "sources": ["ryan_serhant_ai"],
            "powered_by": "Groq (FREE!)"
        }
    }


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), query: Optional[str] = Form(None),
                      experts: Optional[str] = Form(None)):
    """
    Handle file uploads and generate response

    Supports: PDF, TXT, DOCX, XLSX, CSV, JPG, PNG
    experts: JSON array of selected advisor ids, e.g. '["voss","eisen"]'
    """
    try:
        expert_ids = None
        if experts:
            try:
                expert_ids = json.loads(experts)
            except Exception:
                expert_ids = None
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        ext = os.path.splitext(file.filename)[1].lower()
        response_text = None
        file_text = ""

        # Photos (headshots, flyers, listing shots): the panel looks at the ACTUAL image
        if ext in [".jpg", ".jpeg", ".png"]:
            response_text = analyze_image_with_groq(tmp_path, query, file.filename, expert_ids=expert_ids)
            file_text = f"[Image reviewed: {file.filename}]"

        # Documents: extract text (with vision-OCR fallback for scanned PDFs)
        if not response_text:
            file_text = extract_text_from_file(tmp_path, file.filename)
            doc_context = f"[File name: {file.filename}]\n\n{file_text[:24000]}"
            if GROQ_AVAILABLE:
                response_text = await generate_response_with_groq(query or "", "practical",
                                                                  context=doc_context, expert_ids=expert_ids)

        # Last-resort fallback only if Groq is unavailable
        if not response_text:
            response_text = generate_file_response(file_text, query, file.filename)

        # Clean up temp file
        os.unlink(tmp_path)

        return {
            "status": "success",
            "file_name": file.filename,
            "file_size": len(content),
            "extracted_text": file_text[:16000],
            "response": {
                "text": response_text,
                "style": "practical",
                "word_count": len(response_text.split()),
                "reading_time_seconds": len(response_text.split()) // 3,
                "ryan_voice": {
                    "energy": 0.85,
                    "pace_wpm": 145,
                    "confidence": 1.0
                }
            },
            "citations": [
                {
                    "rank": 1,
                    "source": "analysis",
                    "title": f"Analysis of {file.filename}",
                    "quality_score": 85,
                    "excerpt": file_text[:150] + "...",
                    "url": None,
                    "topics": ["document_analysis"]
                }
            ],
            "metadata": {
                "chunks_retrieved": 1,
                "confidence_score": 0.85,
                "generation_latency_ms": 300,
                "sources": ["file_analysis"]
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={"status": "error", "message": f"File processing error: {str(e)}"}
        )


def extract_text_from_file(file_path: str, file_name: str) -> str:
    """Extract text from various file types"""
    ext = os.path.splitext(file_name)[1].lower()

    try:
        if ext == ".txt":
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()

        elif ext == ".pdf":
            text = ""
            try:
                import PyPDF2
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    for page in reader.pages:
                        text += page.extract_text() or ""
            except Exception:
                pass
            # Scanned PDF (no text layer)? OCR it with the free Groq vision model.
            if len(text.strip()) < 200:
                ocr_text = ocr_pdf_with_groq_vision(file_path)
                if ocr_text.strip():
                    return ocr_text
            return text if text.strip() else "[PDF content could not be extracted]"

        elif ext in [".docx", ".doc"]:
            try:
                from docx import Document
                doc = Document(file_path)
                return "\n".join([p.text for p in doc.paragraphs])
            except:
                return "[DOCX extraction requires python-docx. Showing file name: " + file_name + "]"

        elif ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"\n--- Sheet: {sheet} ---\n"
                    for row in ws.iter_rows(values_only=True):
                        text += " | ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                return text if text.strip() else "[Excel file is empty]"
            except:
                return "[Excel extraction requires openpyxl. Showing file name: " + file_name + "]"

        elif ext == ".csv":
            try:
                import csv
                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    text = ""
                    for row in reader:
                        text += " | ".join(row) + "\n"
                return text if text.strip() else "[CSV file is empty]"
            except:
                return "[CSV extraction failed. Showing file name: " + file_name + "]"

        elif ext in [".jpg", ".jpeg", ".png"]:
            return f"[Image file detected: {file_name}. For full OCR, integrate with Tesseract or Claude Vision API]"

        else:
            return f"[Unsupported file type: {ext}]"

    except Exception as e:
        return f"[Error reading file: {str(e)}]"


def generate_file_response(file_text: str, query: Optional[str], file_name: str) -> str:
    """Generate Ryan's response based on file content"""

    # Truncate for demo
    preview = file_text[:500] if len(file_text) > 500 else file_text

    # Detect file type and generate appropriate response
    if "contract" in file_name.lower() or "agreement" in file_name.lower():
        return f"""Looking at this contract/agreement, here's what matters:

First, never sign anything you don't fully understand. Period. I don't care if you're under time pressure—that's when mistakes happen.

Key things to verify:
• Commission structure (what's your take?)
• Contingencies (what kills the deal?)
• Timelines (realistic closing date?)
• Liability (who's responsible if things go wrong?)

From what I see here: "{preview}"

The best move: Have a real estate attorney review this before you commit. It costs $300-500 and saves you thousands in problems down the line.

Remember—in real estate, the deal terms matter as much as the deal itself. Don't be the agent who signed away their commission."""

    elif "listing" in file_name.lower() or "mls" in file_name.lower():
        return f"""Analyzing this listing, here's my take:

This property needs to tell a story that makes buyers FEEL something. Right now, I'm looking at: "{preview}"

What's missing:
• What makes this property special? (The hook)
• Who's the ideal buyer? (Your target)
• What's the emotional benefit? (Not just features)

Here's what I'd do:
1. Rewrite the description to lead with the lifestyle, not the square footage
2. Highlight the most valuable feature (view? location? condition?)
3. Make it clear why THIS home matters

The agents who win are the ones who sell dreams, not properties. Make buyers see themselves living here."""

    elif any(word in file_name.lower() for word in ["proposal", "offer", "bid"]):
        return f"""Looking at this offer/proposal: "{preview}"

Here's the reality: Your first offer sets the tone for the entire negotiation.

Smart moves:
• Lead with confidence (not desperation)
• Include contingencies that protect YOU (not just the buyer)
• Have an expiration date (creates urgency)
• Document everything in writing

What separates top agents from average ones is how they structure the deal. This gives you negotiating room throughout the process.

Bottom line: This offer is solid, but make sure you're not leaving money on the table. What's your walkaway price?"""

    else:
        return f"""I reviewed your document: "{preview}"

Here's my professional take:

The best real estate documents do three things:
1. Protect your interests (always)
2. Clarify expectations (zero ambiguity)
3. Create leverage (for future negotiations)

From what I see, you're on the right track. But remember—in this business, details matter. A single clause can change who makes money and who loses it.

Key question: Did you have this reviewed by someone with experience? Because once you sign, you're committed. Make sure you're comfortable with every line."""


@app.get("/")
async def root():
    """Serve the landing page"""
    with open("landing.html", "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

@app.get("/chat")
async def chat():
    """Serve the chat/coaching interface"""
    with open("frontend.html", "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


@app.get("/admin/dashboard")
async def dashboard():
    """Serve admin dashboard"""
    return HTMLResponse(content=open("backend/admin/dashboard.html", "r", encoding="utf-8").read())


if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*70)
    print("🚀 RYAN SERHANT RESPONSE TOOL - DEMO SERVER")
    print("="*70)
    print("\n📍 API Endpoints:")
    print("   GET  /api/health      → System status")
    print("   GET  /api/status      → Detailed metrics")
    print("   POST /api/query       → Get Ryan response")
    print("\n🌐 Frontend:")
    print("   http://localhost:8000")
    print("\n📊 Admin Dashboard:")
    print("   http://localhost:8000/admin/dashboard")
    print("\n" + "="*70 + "\n")

    uvicorn.run(app, host="0.0.0.0", port=8000)
