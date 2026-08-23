# -*- coding: utf-8 -*-
"""Build the native-speaker review sheet for the Spanish library.

Takes a STRATIFIED sample - every topic x format combination gets covered, and
the idiom-heavy formats (call, text) are favoured, because that is where
translated-sounding language actually costs an agent a conversation.

The reviewer only has to do two things per row: pick a verdict from the
dropdown, and paste a better line if they have one. Their corrections come back
as phrasebook entries, so one red-line improves all 258 scripts on the re-run.

    python scripts/spanish_review_sheet.py            # ~25 rows
    python scripts/spanish_review_sheet.py --rows 40
"""
import argparse
import importlib.util
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
GOLD = "C9A85C"
INK = "131316"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", type=int, default=25)
    ap.add_argument("--out", default=str(ROOT / "Spanish Script Review.xlsx"))
    args = ap.parse_args()

    spec = importlib.util.spec_from_file_location("bs", str(ROOT / "scripts" / "build_spanish.py"))
    bs = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(bs)

    es = json.loads((ROOT / "voices" / "_es.json").read_text(encoding="utf-8"))
    scripts = [s for s in bs.load_scripts() if s["key"] in es]
    if not scripts:
        sys.exit("No Spanish scripts built yet - run scripts/build_spanish.py first.")

    # Stratify properly: every format must be represented, and within a format
    # spread across as many different lead types as possible. Calls and texts get
    # a slightly bigger share - spoken naturalness is where a translated-sounding
    # line actually costs a conversation - but vm and email are never zero.
    weight = {"call": 0.32, "text": 0.28, "vm": 0.20, "email": 0.20}
    by_ch = {}
    for s in scripts:
        by_ch.setdefault(s["ch"], []).append(s)

    sample, used_topics = [], set()
    for ch, share in weight.items():
        pool = by_ch.get(ch, [])
        if not pool:
            continue
        want = max(1, round(args.rows * share))
        # prefer lead types not already covered, so the grid spreads
        pool = sorted(pool, key=lambda s: (s["t"] in used_topics, s["t"]))
        picked, seen_t = [], set()
        for s in pool:
            if s["t"] in seen_t:
                continue
            seen_t.add(s["t"])
            picked.append(s)
            used_topics.add(s["t"])
            if len(picked) >= want:
                break
        sample.extend(picked)

    for s in scripts:                      # top up to the requested row count
        if len(sample) >= args.rows:
            break
        if s not in sample:
            sample.append(s)
    sample = sample[:args.rows]
    order = {"call": 0, "text": 1, "vm": 2, "email": 3}
    sample.sort(key=lambda s: (order.get(s["ch"], 9), s["t"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Spanish Review"

    ws["A1"] = "Spanish Script Review"
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=GOLD)
    ws["A2"] = ("For each script: read the Spanish as if you were saying it to a homeowner in LA. "
                "Pick a verdict, and if it should be worded differently, paste your version. "
                "Industry words left in English (listing, escrow, short sale, forbearance) are intentional.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells("A2:F2")

    heads = ["#", "Lead Type", "Format", "English (original)", "Spanish (review this)",
             "Verdict", "Your better wording / notes"]
    hrow = 4
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=INK)
        c.alignment = Alignment(vertical="center")

    topic_name = {}
    src = (ROOT / "scripts_data.js").read_text(encoding="utf-8")
    import re
    for tid, tname in re.findall(r'\["(\w+)","([^"]+)","', src):
        topic_name.setdefault(tid, tname)
    ch_name = {"call": "Call", "vm": "Voicemail", "text": "Text", "email": "Email"}

    for n, s in enumerate(sample, 1):
        r = hrow + n
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=topic_name.get(s["t"], s["t"]))
        ws.cell(row=r, column=3, value=ch_name.get(s["ch"], s["ch"]))
        ws.cell(row=r, column=4, value=s["body"])
        ws.cell(row=r, column=5, value=es[s["key"]]["script"])
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value="")
        for col in (4, 5, 7):
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 150

    dv = DataValidation(type="list",
                        formula1='"Use as-is,Minor edit,Rewrite"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("F%d:F%d" % (hrow + 1, hrow + len(sample)))

    for col, w in zip(range(1, 8), (5, 22, 11, 62, 62, 14, 40)):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A%d" % (hrow + 1)

    wb.save(args.out)
    grid = sorted({(s["t"], s["ch"]) for s in sample})
    print("wrote %s" % args.out)
    print("  %d scripts, covering %d topic x format combinations" % (len(sample), len(grid)))
    print("  formats: " + ", ".join("%s=%d" % (c, sum(1 for s in sample if s["ch"] == c))
                                    for c in ("call", "text", "vm", "email")))


if __name__ == "__main__":
    main()
